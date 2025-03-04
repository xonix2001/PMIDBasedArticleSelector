import requests
import time, os
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from Bio import Entrez
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from translator import translator

class PubMedDataProcessorBase:
    def __init__(self, email, api_key):
        self.email = email
        Entrez.email = email
        self.api_key = api_key  
        self.input_file_path = None
        self.output_file_path = None

    @staticmethod
    def get_doi_from_pubmed_record(record):
        for article_id in record['PubmedData']['ArticleIdList']:
            if article_id.attributes['IdType'] == 'doi':
                return str(article_id)
        return None

    def fetch_pubmed_info(self, pmid):
        handle = Entrez.efetch(db="pubmed", id=str(pmid), retmode="xml")
        records = Entrez.read(handle)
        handle.close()
        article = records['PubmedArticle'][0]
        title = article['MedlineCitation']['Article']['ArticleTitle']
        pub_date = article['MedlineCitation']['Article']['Journal']['JournalIssue']['PubDate']['Year']
        journal_info = article['MedlineCitation']['Article']['Journal']
        journal = journal_info.get('Title', journal_info['ISOAbbreviation'])
        doi = self.get_doi_from_pubmed_record(article)
        return title, pub_date, journal, doi

    @staticmethod
    def fetch_citation_count(doi):
        url = f"https://api.crossref.org/works/{doi}"
        response = requests.get(url)
        if response.status_code == 200:
            data = response.json()
            citation_count = data['message'].get('is-referenced-by-count', 0)
        else:
            citation_count = "No Citation Data"
        return citation_count

    @staticmethod
    def format_excel(worksheet):
        for cell in worksheet[1]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for col in worksheet.iter_cols(min_row=2, max_col=worksheet.max_column):
            for cell in col:
                if cell.column in (1, 4):
                    worksheet.column_dimensions[get_column_letter(cell.column)].bestFit = True
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                elif cell.column in (2, 3, 7, 8):
                    worksheet.column_dimensions[get_column_letter(cell.column)].width = 60
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                elif cell.column in (5, 6, 9):
                    worksheet.column_dimensions[get_column_letter(cell.column)].width = 20
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def fetch_and_process_pmid(self, pmid):
        title, pub_date, journal, doi = self.fetch_pubmed_info(pmid)
        title_cn = translator(title) if title else "Translation Error"
        citation_count = self.fetch_citation_count(doi) if doi and doi != "No DOI" else "No Citation Data"
        return title, title_cn, pub_date, journal, citation_count

    def process_data(self, input_file_path, output_file_path=None):
        with open(input_file_path, 'r', encoding='utf-8') as file:
            pmid_list = file.read().splitlines()
        data = {
            "PMID": [],
            "Title": [],
            "Title_CN": [],
            "Year": [],
            "Journal": [],
            "Crossref-Cites": []
        }
        with ThreadPoolExecutor(max_workers=10) as executor:
            future_to_pmid = {executor.submit(self.fetch_and_process_pmid, pmid): pmid for pmid in pmid_list}
            for future in as_completed(future_to_pmid):
                pmid = future_to_pmid[future]
                try:
                    result = future.result()
                    data["PMID"].append(pmid)
                    data["Title"].append(result[0])
                    data["Title_CN"].append(result[1])
                    data["Year"].append(result[2])
                    data["Journal"].append(result[3])
                    data["Crossref-Cites"].append(result[4])
                except Exception as e:
                    print(f"Error processing PMID {pmid}: {e}")
                time.sleep(0.01)  # API查询的时间间隔0.01s
        df = pd.DataFrame(data)
        df.sort_values(by=['Year', 'Crossref-Cites'], ascending=[False, False], inplace=True)
        if output_file_path:
            excel_path = output_file_path
        else:
            excel_filename = f"ShortInfo_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            excel_path = os.path.join(os.path.dirname(input_file_path), excel_filename)
        df.to_excel(excel_path, index=False, engine='openpyxl')
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active
        self.format_excel(sheet)
        workbook.save(excel_path)
        print(f"数据处理完成，已保存至 {excel_path}")
        
    @staticmethod
    def add_hyperlink_to_URL(sheet, url_column_index):
        """
        为Excel工作表中指定列的URL添加超链接，不使用workbook属性。
        :param sheet: openpyxl的Worksheet对象。
        :param url_column_index: 包含URL的列的索引，索引从1开始。
        """
        # 定义超链接样式
        hyperlink_font = Font(color="0000FF", underline="single")
        # 遍历指定列中的每个单元格，为URL设置超链接
        for row in range(2, sheet.max_row + 1):  # 从第2行开始遍历，因为第1行是列标题
            cell = sheet.cell(row, url_column_index)
            cell.hyperlink = cell.value  # 设置超链接
            cell.font = hyperlink_font  # 应用超链接样式

    def run(self, input_file_path=None, output_file_path=None):
        if not input_file_path:
            print("未提供输入文件路径。")
            return
        self.input_file_path = input_file_path
        self.output_file_path = output_file_path
        self.process_data(self.input_file_path, self.output_file_path)
        

if __name__ == "__main__":
    email = "263026560@qq.com"
    api_key = "c440fd575a5e0b26840b67cf3280c9b9bf08"
    processor = PubMedDataProcessorBase(email, api_key)
    processor.run(input_file_path="D:\python code\塔吉科研\8\pmids_list\epilepsy [Title Abstract] AND neuroscience [Title .txt", output_file_path="D:\\python code\\塔吉科研\\8\\pmids.xlsx")
